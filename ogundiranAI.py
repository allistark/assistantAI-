import streamlit as st
from groq import Groq
import PyPDF2
import docx
import openpyxl

# ---- SETUP ----
st.set_page_config(
    page_title="LexAI - Legal Assistant",
    page_icon="⚖️",
    layout="wide"
)

# ---- GROQ CONNECTION ----
GROQ_API_KEY = st.secrets["GROQ_API_KEY"]
client = Groq(api_key=GROQ_API_KEY)

# ---- MEMORY ----
if "chat_history" not in st.session_state:
    st.session_state.chat_history = []
if "document_text" not in st.session_state:
    st.session_state.document_text = ""
if "doc_loaded" not in st.session_state:
    st.session_state.doc_loaded = False
if "file_names" not in st.session_state:
    st.session_state.file_names = []

# ---- FILE READERS ----
def read_pdf(file):
    text = ""
    pdf = PyPDF2.PdfReader(file)
    for page in pdf.pages:
        text += page.extract_text() + "\n"
    return text

def read_docx(file):
    text = ""
    doc = docx.Document(file)
    for para in doc.paragraphs:
        text += para.text + "\n"
    return text

def read_excel(file):
    text = ""
    wb = openpyxl.load_workbook(file)
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        text += f"Sheet: {sheet}\n"
        for row in ws.iter_rows(values_only=True):
            row_text = " | ".join(
                [str(c) for c in row if c]
            )
            if row_text:
                text += row_text + "\n"
    return text

# ---- SIDEBAR ----
with st.sidebar:
    st.image(
        "https://img.icons8.com/color/96/scales--v1.png",
        width=80
    )
    st.title("LexAI")
    st.caption("Legal Intelligence Platform")
    st.divider()
    st.subheader("📁 Upload Documents")
    st.caption("Supported: PDF • Word • Excel")
    uploaded_files = st.file_uploader(
        "Choose files",
        type=["pdf", "docx", "xlsx"],
        accept_multiple_files=True,
        label_visibility="collapsed"
    )
    if uploaded_files:
        if st.button(
            "⚡ Process Documents",
            type="primary",
            use_container_width=True
        ):
            with st.spinner("Reading documents..."):
                all_text = ""
                names = []
                for file in uploaded_files:
                    try:
                        ext = file.name.split(".")[-1].lower()
                        if ext == "pdf":
                            text = read_pdf(file)
                        elif ext == "docx":
                            text = read_docx(file)
                        elif ext == "xlsx":
                            text = read_excel(file)
                        all_text += f"\n\n=== {file.name} ===\n"
                        all_text += text
                        names.append(file.name)
                        st.success(f"✅ {file.name}")
                    except:
                        st.error(f"❌ {file.name}")
                st.session_state.document_text = all_text
                st.session_state.doc_loaded = True
                st.session_state.file_names = names
                st.session_state.chat_history = []
    st.divider()
    if st.session_state.file_names:
        st.subheader("📂 Loaded Files")
        for name in st.session_state.file_names:
            st.caption(f"📄 {name}")
        st.divider()
    if st.button(
        "🗑️ Clear Everything",
        use_container_width=True
    ):
        st.session_state.chat_history = []
        st.session_state.document_text = ""
        st.session_state.doc_loaded = False
        st.session_state.file_names = []
        st.rerun()
    st.divider()
    st.caption("⚠️ For informational purposes only.")
    st.caption("Always consult a licensed attorney.")

# ---- MAIN PAGE ----
st.title("⚖️ LexAI — Legal Intelligence Platform")
st.caption("Upload legal documents and get instant AI answers")
st.divider()

# Stats
col1, col2, col3 = st.columns(3)
col1.metric("📄 Documents", len(st.session_state.file_names))
col2.metric("💬 Questions", len(st.session_state.chat_history) // 2)
col3.metric("🟢 Status", "Ready" if st.session_state.doc_loaded else "No Documents")
st.divider()

# ---- CHAT ----
if not st.session_state.doc_loaded:
    st.info("👈 Upload your legal documents on the left to get started")
    col1, col2, col3 = st.columns(3)
    with col1:
        st.success("📄 Step 1\n\nUpload your PDF, Word or Excel legal documents")
    with col2:
        st.success("⚡ Step 2\n\nClick Process Documents to let AI read them")
    with col3:
        st.success("💬 Step 3\n\nAsk any question about your documents")
else:
    for msg in st.session_state.chat_history:
        if msg["role"] == "user":
            with st.chat_message("user"):
                st.write(msg["content"])
        else:
            with st.chat_message(
                "assistant",
                avatar="⚖️"
            ):
                st.write(msg["content"])
    question = st.chat_input(
        "Ask a question about your legal documents..."
    )
    if question:
        with st.chat_message("user"):
            st.write(question)
        system_prompt = """You are LexAI, a professional
        legal AI assistant with 20 years of experience.
        Your rules:
        1. Only answer from the uploaded documents
        2. Be professional and precise
        3. Mention which document the answer came from
        4. If not found say:
           'This was not found in your documents'
        5. Always recommend consulting a licensed
           attorney for final legal decisions"""
        user_prompt = f"""Legal Documents:
{st.session_state.document_text}
Lawyer Question: {question}"""
        with st.chat_message("assistant", avatar="⚖️"):
            with st.spinner("Analyzing documents..."):
                response = client.chat.completions.create(
                    model="llama-3.3-70b-versatile",
                    messages=[
                        {
                            "role": "system",
                            "content": system_prompt
                        },
                        {
                            "role": "user",
                            "content": user_prompt
                        }
                    ],
                    max_tokens=1500
                )
                answer = response.choices[0].message.content
                st.write(answer)
        st.session_state.chat_history.append({
            "role": "user",
            "content": question
        })
        st.session_state.chat_history.append({
            "role": "assistant",
            "content": answer
        })